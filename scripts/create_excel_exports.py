import os
import re
import ast
import tempfile
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpenPyXLImage
from PIL import Image as PILImage

# Paths
BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
EXPORT_DIR = os.path.join(BASE_DIR, "exports")
os.makedirs(EXPORT_DIR, exist_ok=True)

GT1_PATH = os.path.join(BASE_DIR, "data", "ground_truth", "lohmann_rauscher", "Experte1_LR_GroundTruth.csv")
GT2_PATH = os.path.join(BASE_DIR, "data", "ground_truth", "lohmann_rauscher", "Experte2_LR_GroundTruth.csv")

ZERO_LR_PATH = os.path.join(BASE_DIR, "data", "llm_outputs", "zero_shot_lr", "zero_shot_lr_raw.csv")
FEW_LR_PATH = os.path.join(BASE_DIR, "data", "llm_outputs", "few_shot_lr", "few_shot_lr_raw.csv")
TWO_LR_PATH = os.path.join(BASE_DIR, "data", "llm_outputs", "two_stage_lr", "two_stage_lr_raw.csv")

NURSIT_PATH = os.path.join(BASE_DIR, "data", "ground_truth", "allgemeine_verbandsklassen.csv")
ZERO_NURSIT_PATH = os.path.join(BASE_DIR, "data", "llm_outputs", "zero_shot", "zero_shot_raw", "zero_shot_raw.csv")

WUNDBILDER_DIR = os.path.join(BASE_DIR, "data", "wundbilder")

IMAGE_IDS = [f"wunde_{i+1:02d}" for i in range(60)]

def clean_val(val):
    if pd.isna(val) or val is None:
        return "—"
    val_str = str(val).strip()
    if val_str == "" or val_str == "[]" or val_str == "nan":
        return "—"
    return val_str

# Pre-generate image thumbnails to a temp directory
TEMP_THUMB_DIR = tempfile.mkdtemp()

def get_thumbnail_path(img_id, target_width=270, target_height=180):
    num = int(img_id.replace("wunde_", ""))
    orig_file = os.path.join(WUNDBILDER_DIR, f"Bild{num}.jpg")
    if not os.path.exists(orig_file):
        return None
    thumb_file = os.path.join(TEMP_THUMB_DIR, f"{img_id}_large_thumb.jpg")
    if not os.path.exists(thumb_file):
        im = PILImage.open(orig_file)
        im_resized = im.resize((target_width, target_height), PILImage.Resampling.LANCZOS)
        im_resized.save(thumb_file, "JPEG", quality=92)
    return thumb_file

# Style definitions
FONT_HEADER = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FILL_HEADER = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")

FONT_CAT = Font(name="Calibri", size=11, bold=True, color="0F172A")

FILL_EXP1 = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")  # Soft Green
FILL_EXP2 = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")  # Soft Blue
FILL_NURSIT = PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid")# Soft Mint

FILL_ZERO = PatternFill(start_color="FFF7ED", end_color="FFF7ED", fill_type="solid") # Soft Orange
FILL_FEW = PatternFill(start_color="F3E8FF", end_color="F3E8FF", fill_type="solid")  # Soft Purple
FILL_TWO = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")  # Soft Amber

BORDER_THIN = Side(border_style="thin", color="CBD5E1")
BORDER_MEDIUM_BOTTOM = Side(border_style="medium", color="475569")
CELL_BORDER = Border(left=BORDER_THIN, right=BORDER_THIN, top=BORDER_THIN, bottom=BORDER_THIN)
BLOCK_BOTTOM_BORDER = Border(left=BORDER_THIN, right=BORDER_THIN, top=BORDER_THIN, bottom=BORDER_MEDIUM_BOTTOM)

CATEGORIES_LR = [
    ("Wundtyp", "wundtyp", "wundtyp"),
    ("Lokalisation", "lokalisation", "lokalisation"),
    ("Wundstadium", "wundstadium", "wundstadium"),
    ("Wundgrund", "wundgrund", "wundgrund"),
    ("Wundrand", "wundrand", "wundrand"),
    ("Wundumgebung", "wundumgebung", "wundumgebung"),
    ("Exsudat", "exsudat", "exsudat_menge"),
    ("Auffälligkeiten / Anmerkungen", "auffaelligkeiten", "weitere_auffaelligkeiten"),
    ("Debridement Notwendig", "debridement_notwendig", "debridement_notwendig"),
    ("Debridement Methode", "debridement", "debridement_methode"),
    ("Infektionsverdacht", "infektion", "infektion_vorhanden"),
    ("Spüllösung", "spuelloesung", "spuelloesung"),
    ("Primärverband (Präferenz)", "praeferenz_produkt", "praeferenz_wundauflage"),
    ("Primärverband (Alternativ)", "alternative_produkt", "alternativ_wundauflage"),
    ("Sekundärverband (Präferenz)", "ergaenzende_produkte_praeferenz", "praeferenz_ergaenzung"),
    ("Sekundärverband (Alternativ)", "ergaenzende_produkte_alternativ", "alternativ_ergaenzung"),
    ("Kompression Indiziert", "kompression_indiziert", "kompression_indiziert"),
    ("Kompression Produkt", "kompression_produkte", "kompression_produkt"),
    ("Einschränkungen / Annahmen", "einschraenkungen", "einschraenkungen_annahmen")
]

CATEGORIES_NURSIT = [
    ("Wundtyp", "wundtyp", "wundtyp"),
    ("Wundtyp Spezifikation", "wundtyp_spezifikation", "wundtyp_spezifizierung"),
    ("Lokalisation", "lokalisation", "lokalisation"),
    ("Wundstadium", "wundstadium", "wundphase"),
    ("Exsudat", "exsudat", "exsudat_menge"),
    ("Infektion / Verdacht", "infektion", "infektionsstatus"),
    ("Antimikrobielles Agens", "antimikrobielles_agens", "antimikrobielles_agens"),
    ("Antimikrobiell Notwendig", "antimikrobiell_notwendig", "antimikrobieller_verband"),
    ("Wundrand", "wundrand", "wundrand"),
    ("Wundumgebung", "wundumgebung", "wundumgebung"),
    ("Auffälligkeiten", "auffaelligkeiten", "weitere_auffaelligkeiten"),
    ("Debridement Notwendig", "debridement_notwendig", "debridement_notwendig"),
    ("Debridement Methode", "debridement", "debridement_methode"),
    ("Spüllösung", "spuelloesung", "spuelloesung"),
    ("Primärverband (Präferenz)", "praeferenz_produkt", "praeferenz_verbandklasse"),
    ("Primärverband (Alternativ)", "alternative_produkt", "alternativ_verbandklasse"),
    ("Sekundärverband / Fixierung", "sekundaerverband", "sekundaerverband_fixierung"),
    ("Hautschutz", "hautschutz", "wundrand_hautschutz"),
    ("Kompression Indiziert", "kompression_indiziert", "kompression_indiziert"),
    ("Kompression Produkte", "kompression_produkte", "kompression_art"),
    ("Einschränkungen", "einschraenkungen", "einschraenkungen_annahmen")
]

def add_header_and_images(ws):
    headers = ["Kategorie", "Quelle / System"] + IMAGE_IDS
    ws.append(headers)
    
    ws.row_dimensions[1].height = 25
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[2].height = 140
    img_row_cells = ["Wundbild", "Vorschau"] + [""] * 60
    ws.append(img_row_cells)

    ws.cell(row=2, column=1).font = FONT_CAT
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=2, column=2).font = Font(name="Calibri", size=10, bold=True)
    ws.cell(row=2, column=2).alignment = Alignment(horizontal="center", vertical="center")

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.border = BLOCK_BOTTOM_BORDER
        cell.fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

    for idx, img_id in enumerate(IMAGE_IDS):
        col_letter = get_column_letter(idx + 3)
        thumb_path = get_thumbnail_path(img_id, target_width=270, target_height=180)
        if thumb_path:
            img = OpenPyXLImage(thumb_path)
            ws.add_image(img, f"{col_letter}2")

    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 16
    for i in range(3, len(headers) + 1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = 36

def get_dict_from_df(df, col):
    res = {}
    for _, row in df.iterrows():
        img_id = str(row.get("image_id", "")).strip()
        if img_id:
            res[img_id] = clean_val(row.get(col, ""))
    return res

# ----------------------------------------------------
# 1. EXCEL: Lohmann & Rauscher (4 Untertabs)
# ----------------------------------------------------
def build_excel_lr():
    print("Building Excel 1: Lohmann_Rauscher_Vergleich_Experten_KI.xlsx...")
    df_gt1 = pd.read_csv(GT1_PATH, sep=";")
    df_gt2 = pd.read_csv(GT2_PATH, sep=";")
    df_zero = pd.read_csv(ZERO_LR_PATH, sep=",")
    df_few = pd.read_csv(FEW_LR_PATH, sep=",")
    df_two = pd.read_csv(TWO_LR_PATH, sep=",")

    wb = openpyxl.Workbook()

    # --- TAB 1: Alle im Vergleich ---
    ws0 = wb.active
    ws0.title = "Alle im Vergleich"
    ws0.views.sheetView[0].showGridLines = True
    add_header_and_images(ws0)

    r_idx = 3
    for cat_name, gt_col, llm_col in CATEGORIES_LR:
        gt1_dict = get_dict_from_df(df_gt1, gt_col)
        gt2_dict = get_dict_from_df(df_gt2, gt_col)
        zero_dict = get_dict_from_df(df_zero, llm_col)
        few_dict = get_dict_from_df(df_few, llm_col)
        two_dict = get_dict_from_df(df_two, llm_col)

        rows_config = [
            ("Experte 1", gt1_dict, FILL_EXP1),
            ("Experte 2", gt2_dict, FILL_EXP2),
            ("KI Zero-Shot", zero_dict, FILL_ZERO),
            ("KI Few-Shot", few_dict, FILL_FEW),
            ("KI Two-Stage", two_dict, FILL_TWO)
        ]

        start_r = r_idx
        for sys_idx, (sys_name, d_map, fill_style) in enumerate(rows_config):
            row_data = [cat_name if sys_idx == 0 else "", sys_name] + [d_map.get(img_id, "—") for img_id in IMAGE_IDS]
            ws0.append(row_data)

            is_last = (sys_idx == len(rows_config) - 1)
            row_border = BLOCK_BOTTOM_BORDER if is_last else CELL_BORDER

            for c_idx in range(1, len(row_data) + 1):
                cell = ws0.cell(row=r_idx, column=c_idx)
                cell.border = row_border
                cell.fill = fill_style
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.font = FONT_CAT if c_idx == 1 else (Font(name="Calibri", size=10, bold=True) if c_idx == 2 else Font(name="Calibri", size=10))

            r_idx += 1

        end_r = r_idx - 1
        ws0.merge_cells(start_row=start_r, start_column=1, end_row=end_r, end_column=1)
        ws0.cell(row=start_r, column=1).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # --- TAB 2: Experte 1 ---
    ws1 = wb.create_sheet(title="Experte 1")
    ws1.views.sheetView[0].showGridLines = True
    add_header_and_images(ws1)

    r_idx = 3
    for cat_name, gt_col, _ in CATEGORIES_LR:
        val_dict = get_dict_from_df(df_gt1, gt_col)
        row_data = [cat_name, "Experte 1"] + [val_dict.get(img_id, "—") for img_id in IMAGE_IDS]
        ws1.append(row_data)
        for c_idx in range(1, len(row_data) + 1):
            cell = ws1.cell(row=r_idx, column=c_idx)
            cell.border = BLOCK_BOTTOM_BORDER
            cell.fill = FILL_EXP1
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.font = FONT_CAT if c_idx == 1 else (Font(name="Calibri", size=10, bold=True) if c_idx == 2 else Font(name="Calibri", size=10))
        r_idx += 1

    # --- TAB 3: Experte 2 ---
    ws2 = wb.create_sheet(title="Experte 2")
    ws2.views.sheetView[0].showGridLines = True
    add_header_and_images(ws2)

    r_idx = 3
    for cat_name, gt_col, _ in CATEGORIES_LR:
        val_dict = get_dict_from_df(df_gt2, gt_col)
        row_data = [cat_name, "Experte 2"] + [val_dict.get(img_id, "—") for img_id in IMAGE_IDS]
        ws2.append(row_data)
        for c_idx in range(1, len(row_data) + 1):
            cell = ws2.cell(row=r_idx, column=c_idx)
            cell.border = BLOCK_BOTTOM_BORDER
            cell.fill = FILL_EXP2
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.font = FONT_CAT if c_idx == 1 else (Font(name="Calibri", size=10, bold=True) if c_idx == 2 else Font(name="Calibri", size=10))
        r_idx += 1

    # --- TAB 4: KI-Antworten ---
    ws3 = wb.create_sheet(title="KI-Antworten")
    ws3.views.sheetView[0].showGridLines = True
    add_header_and_images(ws3)

    r_idx = 3
    for cat_name, _, llm_col in CATEGORIES_LR:
        zero_dict = get_dict_from_df(df_zero, llm_col)
        few_dict = get_dict_from_df(df_few, llm_col)
        two_dict = get_dict_from_df(df_two, llm_col)

        ki_rows = [
            ("KI Zero-Shot", zero_dict, FILL_ZERO),
            ("KI Few-Shot", few_dict, FILL_FEW),
            ("KI Two-Stage", two_dict, FILL_TWO)
        ]

        start_r = r_idx
        for sys_idx, (sys_name, d_map, fill_style) in enumerate(ki_rows):
            row_data = [cat_name if sys_idx == 0 else "", sys_name] + [d_map.get(img_id, "—") for img_id in IMAGE_IDS]
            ws3.append(row_data)

            is_last = (sys_idx == len(ki_rows) - 1)
            row_border = BLOCK_BOTTOM_BORDER if is_last else CELL_BORDER

            for c_idx in range(1, len(row_data) + 1):
                cell = ws3.cell(row=r_idx, column=c_idx)
                cell.border = row_border
                cell.fill = fill_style
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.font = FONT_CAT if c_idx == 1 else (Font(name="Calibri", size=10, bold=True) if c_idx == 2 else Font(name="Calibri", size=10))

            r_idx += 1

        end_r = r_idx - 1
        ws3.merge_cells(start_row=start_r, start_column=1, end_row=end_r, end_column=1)
        ws3.cell(row=start_r, column=1).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    out_path = os.path.join(EXPORT_DIR, "Lohmann_Rauscher_Vergleich_Experten_KI.xlsx")
    wb.save(out_path)
    print(f"Saved: {out_path}")

# ----------------------------------------------------
# 2. EXCEL: NursIT Allgemeine Verbandsklassen (3 Untertabs)
# ----------------------------------------------------
def build_excel_nursit():
    print("Building Excel 2: NursIT_Allgemeine_Verbandsklassen.xlsx...")
    df_nursit = pd.read_csv(NURSIT_PATH)
    df_zero_nursit = pd.read_csv(ZERO_NURSIT_PATH, sep=",")
    
    df_zero_lr = pd.read_csv(ZERO_LR_PATH, sep=",")
    df_few_lr = pd.read_csv(FEW_LR_PATH, sep=",")
    df_two_lr = pd.read_csv(TWO_LR_PATH, sep=",")

    wb = openpyxl.Workbook()

    # --- TAB 1: Alle im Vergleich ---
    ws0 = wb.active
    ws0.title = "Alle im Vergleich"
    ws0.views.sheetView[0].showGridLines = True
    add_header_and_images(ws0)

    r_idx = 3
    for cat_name, gt_col, llm_col in CATEGORIES_NURSIT:
        nursit_dict = get_dict_from_df(df_nursit, gt_col)
        z_dict = get_dict_from_df(df_zero_nursit if llm_col in df_zero_nursit.columns else df_zero_lr, llm_col if llm_col in df_zero_nursit.columns else (gt_col if gt_col in df_zero_lr.columns else llm_col))
        f_dict = get_dict_from_df(df_few_lr, gt_col if gt_col in df_few_lr.columns else llm_col)
        t_dict = get_dict_from_df(df_two_lr, gt_col if gt_col in df_two_lr.columns else llm_col)

        rows_config = [
            ("NursIT Experte", nursit_dict, FILL_NURSIT),
            ("KI Zero-Shot", z_dict, FILL_ZERO),
            ("KI Few-Shot", f_dict, FILL_FEW),
            ("KI Two-Stage", t_dict, FILL_TWO)
        ]

        start_r = r_idx
        for sys_idx, (sys_name, d_map, fill_style) in enumerate(rows_config):
            row_data = [cat_name if sys_idx == 0 else "", sys_name] + [d_map.get(img_id, "—") for img_id in IMAGE_IDS]
            ws0.append(row_data)

            is_last = (sys_idx == len(rows_config) - 1)
            row_border = BLOCK_BOTTOM_BORDER if is_last else CELL_BORDER

            for c_idx in range(1, len(row_data) + 1):
                cell = ws0.cell(row=r_idx, column=c_idx)
                cell.border = row_border
                cell.fill = fill_style
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.font = FONT_CAT if c_idx == 1 else (Font(name="Calibri", size=10, bold=True) if c_idx == 2 else Font(name="Calibri", size=10))

            r_idx += 1

        end_r = r_idx - 1
        ws0.merge_cells(start_row=start_r, start_column=1, end_row=end_r, end_column=1)
        ws0.cell(row=start_r, column=1).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # --- TAB 2: NursIT Experte ---
    ws1 = wb.create_sheet(title="NursIT Experte")
    ws1.views.sheetView[0].showGridLines = True
    add_header_and_images(ws1)

    r_idx = 3
    for cat_name, gt_col, _ in CATEGORIES_NURSIT:
        val_dict = get_dict_from_df(df_nursit, gt_col)
        row_data = [cat_name, "NursIT Experte"] + [val_dict.get(img_id, "—") for img_id in IMAGE_IDS]
        ws1.append(row_data)
        for c_idx in range(1, len(row_data) + 1):
            cell = ws1.cell(row=r_idx, column=c_idx)
            cell.border = BLOCK_BOTTOM_BORDER
            cell.fill = FILL_NURSIT
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.font = FONT_CAT if c_idx == 1 else (Font(name="Calibri", size=10, bold=True) if c_idx == 2 else Font(name="Calibri", size=10))
        r_idx += 1

    # --- TAB 3: KI-Antworten ---
    ws2 = wb.create_sheet(title="KI-Antworten")
    ws2.views.sheetView[0].showGridLines = True
    add_header_and_images(ws2)

    r_idx = 3
    for cat_name, gt_col, llm_col in CATEGORIES_NURSIT:
        z_dict = get_dict_from_df(df_zero_nursit if llm_col in df_zero_nursit.columns else df_zero_lr, llm_col if llm_col in df_zero_nursit.columns else (gt_col if gt_col in df_zero_lr.columns else llm_col))
        f_dict = get_dict_from_df(df_few_lr, gt_col if gt_col in df_few_lr.columns else llm_col)
        t_dict = get_dict_from_df(df_two_lr, gt_col if gt_col in df_two_lr.columns else llm_col)

        ki_rows = [
            ("KI Zero-Shot", z_dict, FILL_ZERO),
            ("KI Few-Shot", f_dict, FILL_FEW),
            ("KI Two-Stage", t_dict, FILL_TWO)
        ]

        start_r = r_idx
        for sys_idx, (sys_name, d_map, fill_style) in enumerate(ki_rows):
            row_data = [cat_name if sys_idx == 0 else "", sys_name] + [d_map.get(img_id, "—") for img_id in IMAGE_IDS]
            ws2.append(row_data)

            is_last = (sys_idx == len(ki_rows) - 1)
            row_border = BLOCK_BOTTOM_BORDER if is_last else CELL_BORDER

            for c_idx in range(1, len(row_data) + 1):
                cell = ws2.cell(row=r_idx, column=c_idx)
                cell.border = row_border
                cell.fill = fill_style
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.font = FONT_CAT if c_idx == 1 else (Font(name="Calibri", size=10, bold=True) if c_idx == 2 else Font(name="Calibri", size=10))

            r_idx += 1

        end_r = r_idx - 1
        ws2.merge_cells(start_row=start_r, start_column=1, end_row=end_r, end_column=1)
        ws2.cell(row=start_r, column=1).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    out_path = os.path.join(EXPORT_DIR, "NursIT_Allgemeine_Verbandsklassen.xlsx")
    wb.save(out_path)
    print(f"Saved: {out_path}")

if __name__ == "__main__":
    build_excel_lr()
    build_excel_nursit()
