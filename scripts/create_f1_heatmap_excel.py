import os
import ast
import sys
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Ensure export directory exists
EXPORT_DIR = os.path.join(os.path.dirname(__file__), "..", "exports")
os.makedirs(EXPORT_DIR, exist_ok=True)

BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
if BASE_DIR not in sys.path:
    sys.path.insert(0, BASE_DIR)

# File paths
GT1_PATH = os.path.join(BASE_DIR, "data", "ground_truth", "lohmann_rauscher", "Experte1_LR_GroundTruth_normalised.csv")
GT2_PATH = os.path.join(BASE_DIR, "data", "ground_truth", "lohmann_rauscher", "Experte2_LR_GroundTruth_normalised.csv")

ZERO_PATH = os.path.join(BASE_DIR, "data", "llm_outputs", "zero_shot_lr", "zero_shot_lr_normalised.csv")
FEW_PATH = os.path.join(BASE_DIR, "data", "llm_outputs", "few_shot_lr", "few_shot_lr_normalised.csv")
TWO_PATH = os.path.join(BASE_DIR, "data", "llm_outputs", "two_stage_lr", "two_stage_lr_normalised.csv")

from notebooks.analyse_categories.category_loader import PRODUCT_FAMILY_MAP

def parse_set(val):
    if pd.isna(val) or val == "" or val == "[]" or val == "nan" or val is None:
        return set()
    val_str = str(val).strip()
    if val_str.startswith("["):
        try:
            return set(ast.literal_eval(val_str))
        except:
            pass
    return {val_str}

def set_f1(pred_set: set, gt_set: set) -> float:
    if not pred_set and not gt_set:
        return 1.0
    if not pred_set or not gt_set:
        return 0.0
    intersection = len(pred_set.intersection(gt_set))
    if intersection == 0:
        return 0.0
    precision = intersection / len(pred_set)
    recall = intersection / len(gt_set)
    return 2 * (precision * recall) / (precision + recall)

def best_path_f1(pred_pref: set, pred_alt: set, gt_pref: set, gt_alt: set) -> float:
    cand_p = [pred_pref]
    if pred_alt:
        cand_p.append(pred_alt)
        
    cand_g = [gt_pref]
    if gt_alt:
        cand_g.append(gt_alt)
        
    best_score = 0.0
    for cp in cand_p:
        for cg in cand_g:
            score = set_f1(cp, cg)
            if score > best_score:
                best_score = score
    return best_score

# Product Mappings for Levels 2 and 3
def map_level_2(prod_set):
    if not prod_set:
        return set()
    mapped = set()
    for item in prod_set:
        item_lower = item.strip().lower()
        if item_lower in PRODUCT_FAMILY_MAP:
            mapped.add(PRODUCT_FAMILY_MAP[item_lower])
        else:
            if "suprasorb a" in item_lower:
                mapped.add("Suprasorb A (Alginat)")
            elif "suprasorb p" in item_lower:
                mapped.add("Suprasorb P (Schaumstoff)")
            elif "suprasorb x" in item_lower:
                mapped.add("Suprasorb X (Hydrobalance)")
            elif "suprasorb liquacel" in item_lower:
                mapped.add("Suprasorb Liquacel (Hydrofiber)")
            elif "vliwasorb" in item_lower:
                mapped.add("Vliwasorb (Superabsorber)")
            elif "vliwazell" in item_lower:
                mapped.add("Vliwazell (Saugkompresse)")
            elif "solvaline" in item_lower:
                mapped.add("Solvaline (Folien-Saugkompresse)")
            elif "lomatuell" in item_lower:
                mapped.add("Lomatuell (Gittertüll)")
            elif "vliwaktiv" in item_lower:
                mapped.add("Vliwaktiv (Aktivkohle)")
            elif "suprasorb g" in item_lower or "gel-kompresse" in item_lower or "amorphes gel" in item_lower:
                mapped.add("Suprasorb G (Gel)")
            elif "suprasorb h" in item_lower:
                mapped.add("Suprasorb H (Hydrokolloid)")
            elif "suprasorb f" in item_lower:
                mapped.add("Suprasorb F (Folie)")
            elif "cnp" in item_lower:
                mapped.add("Suprasorb CNP (NPWT)")
            elif "metalline" in item_lower:
                mapped.add("Metalline")
            else:
                mapped.add(item)
    return mapped

def map_level_3(prod_set):
    if not prod_set:
        return set()
    mapped = set()
    for item in prod_set:
        item_lower = item.strip().lower()
        if "suprasorb a" in item_lower:
            mapped.add("Alginate")
        elif "suprasorb p" in item_lower:
            mapped.add("Schaumstoffverbände (Foam)")
        elif "suprasorb x" in item_lower:
            mapped.add("Wundkontaktschichten (Silikon/Paraffin)")
        elif "suprasorb liquacel" in item_lower:
            mapped.add("Hydrofaser / Hydrofiber")
        elif "vliwasorb" in item_lower or "vliwazell" in item_lower:
            mapped.add("Superabsorber")
        elif "solvaline" in item_lower or "lomatuell" in item_lower:
            mapped.add("Wundkontaktschichten (Silikon/Paraffin)")
        elif "vliwaktiv" in item_lower:
            mapped.add("Aktivkohleverband")
        elif "suprasorb g" in item_lower or "gel-kompresse" in item_lower or "amorphes gel" in item_lower:
            mapped.add("Hydrogele (Kompresse)")
        elif "suprasorb h" in item_lower:
            mapped.add("Hydrokolloide")
        elif "suprasorb f" in item_lower:
            mapped.add("Semipermeable Filme")
        elif "cnp" in item_lower:
            mapped.add("NPWT / Unterdrucktherapie")
        elif "metalline" in item_lower:
            mapped.add("Wundkontaktschichten (Silikon/Paraffin)")
        else:
            mapped.add("Sonstige Primärverbände")
    return mapped

# Colors & Fills
FILL_DARK_HEADER = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid") # Dark Blue Header
FILL_SECTION_BLUE = PatternFill(start_color="93C5FD", end_color="93C5FD", fill_type="solid") # Blue Level Banner
FILL_SUBSECTION_GREEN = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid") # Light Green Banner
FILL_INTERRATER = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid") # Soft Gray-Blue for Inter-Rater Label

# Heatmap Fills
FILL_GRAY = PatternFill(start_color="B0BEC5", end_color="B0BEC5", fill_type="solid") # Empty Expert Set
FILL_RED = PatternFill(start_color="E57373", end_color="E57373", fill_type="solid") # F1 = 0.0
FILL_ORANGE = PatternFill(start_color="FFB74D", end_color="FFB74D", fill_type="solid") # 0.0 < F1 < 0.5
FILL_YELLOW = PatternFill(start_color="FFF176", end_color="FFF176", fill_type="solid") # F1 = 0.5
FILL_LIGHT_GREEN = PatternFill(start_color="AED581", end_color="AED581", fill_type="solid") # 0.5 < F1 < 1.0
FILL_DARK_GREEN = PatternFill(start_color="66BB6A", end_color="66BB6A", fill_type="solid") # F1 = 1.0

FONT_DARK_HEADER = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
FONT_SECTION = Font(name="Calibri", size=12, bold=True, color="0F172A")
FONT_SUBSECTION = Font(name="Calibri", size=11, bold=True, color="065F46")
FONT_INTERRATER = Font(name="Calibri", size=10, bold=True, color="1E293B")
FONT_ROW_LABEL = Font(name="Calibri", size=10)

BORDER_THIN = Side(border_style="thin", color="CBD5E1")
CELL_BORDER = Border(left=BORDER_THIN, right=BORDER_THIN, top=BORDER_THIN, bottom=BORDER_THIN)

IMAGE_IDS = [f"wunde_{i+1:02d}" for i in range(60)]

def get_fill_and_val_str(val):
    if val is None:
        return FILL_GRAY, ""
    val_round = round(val, 4)
    if val_round == 0.0:
        fill = FILL_RED
    elif 0.0 < val_round < 0.5:
        fill = FILL_ORANGE
    elif val_round == 0.5:
        fill = FILL_YELLOW
    elif 0.5 < val_round < 1.0:
        fill = FILL_LIGHT_GREEN
    else: # 1.0
        fill = FILL_DARK_GREEN

    # Format string (0, 0,3, 0,5, 0,7, 1)
    if val_round == 0.0:
        val_str = "0"
    elif val_round == 1.0:
        val_str = "1"
    elif val_round == 0.5:
        val_str = "0,5"
    else:
        val_str = f"{val:.1f}".replace(".", ",")
        
    return fill, val_str

def build_excel_heatmap():
    print("Loading datasets...")
    df_gt1 = pd.read_csv(GT1_PATH, sep=";")
    df_gt2 = pd.read_csv(GT2_PATH, sep=";")
    df_zero = pd.read_csv(ZERO_PATH, sep=",")
    df_few = pd.read_csv(FEW_PATH, sep=",")
    df_two = pd.read_csv(TWO_PATH, sep=",")

    wb = openpyxl.Workbook()

    levels = [
        ("Produkt Ebene", 1),
        ("Unterkategorie Ebene", 2),
        ("Allgemeine Verbandsklassen Ebene", 3)
    ]
    
    # ---------------------------------------------------------
    # SHEET 1: F1 Heatmap (Maximal Kompakt - KI vs Exp 1 & 2)
    # ---------------------------------------------------------
    ws0 = wb.active
    ws0.title = "F1 Heatmap (Maximal Kompakt)"
    ws0.views.sheetView[0].showGridLines = True

    ws0.row_dimensions[1].height = 100
    ws0.column_dimensions['A'].width = 36

    for i, img_id in enumerate(IMAGE_IDS):
        col_letter = get_column_letter(i + 2)
        ws0.column_dimensions[col_letter].width = 4.5
        
        cell = ws0.cell(row=1, column=i+2, value=img_id)
        cell.font = FONT_DARK_HEADER
        cell.fill = FILL_DARK_HEADER
        cell.alignment = Alignment(textRotation=90, horizontal="center", vertical="center")
        cell.border = CELL_BORDER

    cell_a0 = ws0.cell(row=1, column=1, value="")
    cell_a0.fill = FILL_DARK_HEADER
    cell_a0.border = CELL_BORDER

    current_row = 2

    for level_title, level_num in levels:
        ws0.row_dimensions[current_row].height = 24
        cell_sec = ws0.cell(row=current_row, column=1, value=level_title)
        cell_sec.font = FONT_SECTION
        cell_sec.fill = FILL_SECTION_BLUE
        cell_sec.alignment = Alignment(vertical="center")
        
        for c in range(2, 62):
            c_cell = ws0.cell(row=current_row, column=c)
            c_cell.fill = FILL_SECTION_BLUE
            c_cell.border = CELL_BORDER
        current_row += 1

        # 1. Experten Inter Rater Wert Row
        ws0.row_dimensions[current_row].height = 20
        lbl_inter = ws0.cell(row=current_row, column=1, value="  Experten Inter-Rater Wert")
        lbl_inter.font = FONT_INTERRATER
        lbl_inter.fill = FILL_INTERRATER
        lbl_inter.alignment = Alignment(vertical="center")
        lbl_inter.border = CELL_BORDER

        for img_idx, img_id in enumerate(IMAGE_IDS):
            col_num = img_idx + 2
            
            e1_p_raw = parse_set(df_gt1[df_gt1["image_id"] == img_id]["praeferenz_produkt"].values[0] if len(df_gt1[df_gt1["image_id"] == img_id]) > 0 else "")
            e1_a_raw = parse_set(df_gt1[df_gt1["image_id"] == img_id]["alternative_produkt"].values[0] if len(df_gt1[df_gt1["image_id"] == img_id]) > 0 else "")

            e2_p_raw = parse_set(df_gt2[df_gt2["image_id"] == img_id]["praeferenz_produkt"].values[0] if len(df_gt2[df_gt2["image_id"] == img_id]) > 0 else "")
            e2_a_raw = parse_set(df_gt2[df_gt2["image_id"] == img_id]["alternative_produkt"].values[0] if len(df_gt2[df_gt2["image_id"] == img_id]) > 0 else "")

            if level_num == 1:
                e1_p, e1_a = e1_p_raw, e1_a_raw
                e2_p, e2_a = e2_p_raw, e2_a_raw
            elif level_num == 2:
                e1_p, e1_a = map_level_2(e1_p_raw), map_level_2(e1_a_raw)
                e2_p, e2_a = map_level_2(e2_p_raw), map_level_2(e2_a_raw)
            else:
                e1_p, e1_a = map_level_3(e1_p_raw), map_level_3(e1_a_raw)
                e2_p, e2_a = map_level_3(e2_p_raw), map_level_3(e2_a_raw)

            if (not e1_p and not e1_a) or (not e2_p and not e2_a):
                f1_val = None
            else:
                f1_val = best_path_f1(e1_p, e1_a, e2_p, e2_a)

            fill_style, val_str = get_fill_and_val_str(f1_val)
            data_cell = ws0.cell(row=current_row, column=col_num, value=val_str)
            data_cell.fill = fill_style
            data_cell.border = CELL_BORDER
            data_cell.alignment = Alignment(horizontal="center", vertical="center")
            data_cell.font = Font(name="Calibri", size=9, bold=(f1_val == 1.0 or f1_val == 0.0))

        current_row += 1

        # 3 Rows for KI vs Experte 1 & 2
        ki_rows = [
            ("KI Zero-Shot → Experte 1 und 2", "zero"),
            ("KI Few-Shot → Experte 1 und 2", "few"),
            ("KI Two-Stage → Experte 1 und 2", "two")
        ]

        for row_label, comp_key in ki_rows:
            ws0.row_dimensions[current_row].height = 19
            lbl_cell = ws0.cell(row=current_row, column=1, value=f"  {row_label}")
            lbl_cell.font = FONT_ROW_LABEL
            lbl_cell.border = CELL_BORDER

            for img_idx, img_id in enumerate(IMAGE_IDS):
                col_num = img_idx + 2
                
                e1_p_raw = parse_set(df_gt1[df_gt1["image_id"] == img_id]["praeferenz_produkt"].values[0] if len(df_gt1[df_gt1["image_id"] == img_id]) > 0 else "")
                e1_a_raw = parse_set(df_gt1[df_gt1["image_id"] == img_id]["alternative_produkt"].values[0] if len(df_gt1[df_gt1["image_id"] == img_id]) > 0 else "")

                e2_p_raw = parse_set(df_gt2[df_gt2["image_id"] == img_id]["praeferenz_produkt"].values[0] if len(df_gt2[df_gt2["image_id"] == img_id]) > 0 else "")
                e2_a_raw = parse_set(df_gt2[df_gt2["image_id"] == img_id]["alternative_produkt"].values[0] if len(df_gt2[df_gt2["image_id"] == img_id]) > 0 else "")

                z_p_raw = parse_set(df_zero[df_zero["image_id"] == img_id]["praeferenz_wundauflage"].values[0] if len(df_zero[df_zero["image_id"] == img_id]) > 0 else "")
                z_a_raw = parse_set(df_zero[df_zero["image_id"] == img_id]["alternativ_wundauflage"].values[0] if len(df_zero[df_zero["image_id"] == img_id]) > 0 else "")

                f_p_raw = parse_set(df_few[df_few["image_id"] == img_id]["praeferenz_wundauflage"].values[0] if len(df_few[df_few["image_id"] == img_id]) > 0 else "")
                f_a_raw = parse_set(df_few[df_few["image_id"] == img_id]["alternativ_wundauflage"].values[0] if len(df_few[df_few["image_id"] == img_id]) > 0 else "")

                t_p_raw = parse_set(df_two[df_two["image_id"] == img_id]["praeferenz_wundauflage"].values[0] if len(df_two[df_two["image_id"] == img_id]) > 0 else "")
                t_a_raw = parse_set(df_two[df_two["image_id"] == img_id]["alternativ_wundauflage"].values[0] if len(df_two[df_two["image_id"] == img_id]) > 0 else "")

                if level_num == 1:
                    e1_p, e1_a = e1_p_raw, e1_a_raw
                    e2_p, e2_a = e2_p_raw, e2_a_raw
                    z_p, z_a = z_p_raw, z_a_raw
                    f_p, f_a = f_p_raw, f_a_raw
                    t_p, t_a = t_p_raw, t_a_raw
                elif level_num == 2:
                    e1_p, e1_a = map_level_2(e1_p_raw), map_level_2(e1_a_raw)
                    e2_p, e2_a = map_level_2(e2_p_raw), map_level_2(e2_a_raw)
                    z_p, z_a = map_level_2(z_p_raw), map_level_2(z_a_raw)
                    f_p, f_a = map_level_2(f_p_raw), map_level_2(f_a_raw)
                    t_p, t_a = map_level_2(t_p_raw), map_level_2(t_a_raw)
                else:
                    e1_p, e1_a = map_level_3(e1_p_raw), map_level_3(e1_a_raw)
                    e2_p, e2_a = map_level_3(e2_p_raw), map_level_3(e2_a_raw)
                    z_p, z_a = map_level_3(z_p_raw), map_level_3(z_a_raw)
                    f_p, f_a = map_level_3(f_p_raw), map_level_3(f_a_raw)
                    t_p, t_a = map_level_3(t_p_raw), map_level_3(t_a_raw)

                if comp_key == "zero":
                    ki_p, ki_a = z_p, z_a
                elif comp_key == "few":
                    ki_p, ki_a = f_p, f_a
                else:
                    ki_p, ki_a = t_p, t_a

                e1_empty = (not e1_p and not e1_a)
                e2_empty = (not e2_p and not e2_a)
                ki_empty = (not ki_p and not ki_a)

                if (e1_empty and e2_empty) or ki_empty:
                    f1_val = None
                else:
                    cand_scores = []
                    if not e1_empty:
                        cand_scores.append(best_path_f1(ki_p, ki_a, e1_p, e1_a))
                    if not e2_empty:
                        cand_scores.append(best_path_f1(ki_p, ki_a, e2_p, e2_a))
                    f1_val = max(cand_scores)

                fill_style, val_str = get_fill_and_val_str(f1_val)

                data_cell = ws0.cell(row=current_row, column=col_num, value=val_str)
                data_cell.fill = fill_style
                data_cell.border = CELL_BORDER
                data_cell.alignment = Alignment(horizontal="center", vertical="center")
                data_cell.font = Font(name="Calibri", size=9, bold=(f1_val == 1.0 or f1_val == 0.0))

            current_row += 1

    # ---------------------------------------------------------
    # SHEET 2: F1 Heatmap (Kompakt)
    # ---------------------------------------------------------
    ws1 = wb.create_sheet(title="F1 Heatmap (Kompakt)")
    ws1.views.sheetView[0].showGridLines = True

    ws1.row_dimensions[1].height = 100
    ws1.column_dimensions['A'].width = 36

    for i, img_id in enumerate(IMAGE_IDS):
        col_letter = get_column_letter(i + 2)
        ws1.column_dimensions[col_letter].width = 4.5
        
        cell = ws1.cell(row=1, column=i+2, value=img_id)
        cell.font = FONT_DARK_HEADER
        cell.fill = FILL_DARK_HEADER
        cell.alignment = Alignment(textRotation=90, horizontal="center", vertical="center")
        cell.border = CELL_BORDER

    cell_a1 = ws1.cell(row=1, column=1, value="")
    cell_a1.fill = FILL_DARK_HEADER
    cell_a1.border = CELL_BORDER

    current_row = 2

    for level_title, level_num in levels:
        ws1.row_dimensions[current_row].height = 24
        cell_sec = ws1.cell(row=current_row, column=1, value=level_title)
        cell_sec.font = FONT_SECTION
        cell_sec.fill = FILL_SECTION_BLUE
        cell_sec.alignment = Alignment(vertical="center")
        
        for c in range(2, 62):
            c_cell = ws1.cell(row=current_row, column=c)
            c_cell.fill = FILL_SECTION_BLUE
            c_cell.border = CELL_BORDER
        current_row += 1

        ws1.row_dimensions[current_row].height = 20
        lbl_inter = ws1.cell(row=current_row, column=1, value="  Experten Inter-Rater Wert")
        lbl_inter.font = FONT_INTERRATER
        lbl_inter.fill = FILL_INTERRATER
        lbl_inter.alignment = Alignment(vertical="center")
        lbl_inter.border = CELL_BORDER

        for img_idx, img_id in enumerate(IMAGE_IDS):
            col_num = img_idx + 2
            
            e1_p_raw = parse_set(df_gt1[df_gt1["image_id"] == img_id]["praeferenz_produkt"].values[0] if len(df_gt1[df_gt1["image_id"] == img_id]) > 0 else "")
            e1_a_raw = parse_set(df_gt1[df_gt1["image_id"] == img_id]["alternative_produkt"].values[0] if len(df_gt1[df_gt1["image_id"] == img_id]) > 0 else "")

            e2_p_raw = parse_set(df_gt2[df_gt2["image_id"] == img_id]["praeferenz_produkt"].values[0] if len(df_gt2[df_gt2["image_id"] == img_id]) > 0 else "")
            e2_a_raw = parse_set(df_gt2[df_gt2["image_id"] == img_id]["alternative_produkt"].values[0] if len(df_gt2[df_gt2["image_id"] == img_id]) > 0 else "")

            if level_num == 1:
                e1_p, e1_a = e1_p_raw, e1_a_raw
                e2_p, e2_a = e2_p_raw, e2_a_raw
            elif level_num == 2:
                e1_p, e1_a = map_level_2(e1_p_raw), map_level_2(e1_a_raw)
                e2_p, e2_a = map_level_2(e2_p_raw), map_level_2(e2_a_raw)
            else:
                e1_p, e1_a = map_level_3(e1_p_raw), map_level_3(e1_a_raw)
                e2_p, e2_a = map_level_3(e2_p_raw), map_level_3(e2_a_raw)

            if (not e1_p and not e1_a) or (not e2_p and not e2_a):
                f1_val = None
            else:
                f1_val = best_path_f1(e1_p, e1_a, e2_p, e2_a)

            fill_style, val_str = get_fill_and_val_str(f1_val)
            data_cell = ws1.cell(row=current_row, column=col_num, value=val_str)
            data_cell.fill = fill_style
            data_cell.border = CELL_BORDER
            data_cell.alignment = Alignment(horizontal="center", vertical="center")
            data_cell.font = Font(name="Calibri", size=9, bold=(f1_val == 1.0 or f1_val == 0.0))

        current_row += 1

        exp_blocks = [
            ("Experte 1", 1),
            ("Experte 2", 2)
        ]

        for exp_title, exp_target in exp_blocks:
            ws1.row_dimensions[current_row].height = 20
            cell_sub = ws1.cell(row=current_row, column=1, value=f"  {exp_title}")
            cell_sub.font = FONT_SUBSECTION
            cell_sub.fill = FILL_SUBSECTION_GREEN
            cell_sub.alignment = Alignment(vertical="center")

            for c in range(2, 62):
                c_cell = ws1.cell(row=current_row, column=c)
                c_cell.fill = FILL_SUBSECTION_GREEN
                c_cell.border = CELL_BORDER
            current_row += 1

            if exp_target == 1:
                row_specs = [
                    ("KI Zero-Shot → Experte 1", "zero"),
                    ("KI Few-Shot → Experte 1", "few"),
                    ("KI Two-Stage → Experte 1", "two")
                ]
            else:
                row_specs = [
                    ("KI Zero-Shot → Experte 2", "zero"),
                    ("KI Few-Shot → Experte 2", "few"),
                    ("KI Two-Stage → Experte 2", "two")
                ]

            for row_label, comp_key in row_specs:
                ws1.row_dimensions[current_row].height = 19
                lbl_cell = ws1.cell(row=current_row, column=1, value=f"  {row_label}")
                lbl_cell.font = FONT_ROW_LABEL
                lbl_cell.border = CELL_BORDER

                for img_idx, img_id in enumerate(IMAGE_IDS):
                    col_num = img_idx + 2
                    
                    e1_p_raw = parse_set(df_gt1[df_gt1["image_id"] == img_id]["praeferenz_produkt"].values[0] if len(df_gt1[df_gt1["image_id"] == img_id]) > 0 else "")
                    e1_a_raw = parse_set(df_gt1[df_gt1["image_id"] == img_id]["alternative_produkt"].values[0] if len(df_gt1[df_gt1["image_id"] == img_id]) > 0 else "")

                    e2_p_raw = parse_set(df_gt2[df_gt2["image_id"] == img_id]["praeferenz_produkt"].values[0] if len(df_gt2[df_gt2["image_id"] == img_id]) > 0 else "")
                    e2_a_raw = parse_set(df_gt2[df_gt2["image_id"] == img_id]["alternative_produkt"].values[0] if len(df_gt2[df_gt2["image_id"] == img_id]) > 0 else "")

                    z_p_raw = parse_set(df_zero[df_zero["image_id"] == img_id]["praeferenz_wundauflage"].values[0] if len(df_zero[df_zero["image_id"] == img_id]) > 0 else "")
                    z_a_raw = parse_set(df_zero[df_zero["image_id"] == img_id]["alternativ_wundauflage"].values[0] if len(df_zero[df_zero["image_id"] == img_id]) > 0 else "")

                    f_p_raw = parse_set(df_few[df_few["image_id"] == img_id]["praeferenz_wundauflage"].values[0] if len(df_few[df_few["image_id"] == img_id]) > 0 else "")
                    f_a_raw = parse_set(df_few[df_few["image_id"] == img_id]["alternativ_wundauflage"].values[0] if len(df_few[df_few["image_id"] == img_id]) > 0 else "")

                    t_p_raw = parse_set(df_two[df_two["image_id"] == img_id]["praeferenz_wundauflage"].values[0] if len(df_two[df_two["image_id"] == img_id]) > 0 else "")
                    t_a_raw = parse_set(df_two[df_two["image_id"] == img_id]["alternativ_wundauflage"].values[0] if len(df_two[df_two["image_id"] == img_id]) > 0 else "")

                    if level_num == 1:
                        e1_p, e1_a = e1_p_raw, e1_a_raw
                        e2_p, e2_a = e2_p_raw, e2_a_raw
                        z_p, z_a = z_p_raw, z_a_raw
                        f_p, f_a = f_p_raw, f_a_raw
                        t_p, t_a = t_p_raw, t_a_raw
                    elif level_num == 2:
                        e1_p, e1_a = map_level_2(e1_p_raw), map_level_2(e1_a_raw)
                        e2_p, e2_a = map_level_2(e2_p_raw), map_level_2(e2_a_raw)
                        z_p, z_a = map_level_2(z_p_raw), map_level_2(z_a_raw)
                        f_p, f_a = map_level_2(f_p_raw), map_level_2(f_a_raw)
                        t_p, t_a = map_level_2(t_p_raw), map_level_2(t_a_raw)
                    else:
                        e1_p, e1_a = map_level_3(e1_p_raw), map_level_3(e1_a_raw)
                        e2_p, e2_a = map_level_3(e2_p_raw), map_level_3(e2_a_raw)
                        z_p, z_a = map_level_3(z_p_raw), map_level_3(z_a_raw)
                        f_p, f_a = map_level_3(f_p_raw), map_level_3(f_a_raw)
                        t_p, t_a = map_level_3(t_p_raw), map_level_3(t_a_raw)

                    is_gray = False
                    if exp_target == 1:
                        if not e1_p and not e1_a:
                            is_gray = True
                        elif comp_key == "zero" and (not z_p and not z_a):
                            is_gray = True
                        elif comp_key == "few" and (not f_p and not f_a):
                            is_gray = True
                        elif comp_key == "two" and (not t_p and not t_a):
                            is_gray = True
                    else:
                        if not e2_p and not e2_a:
                            is_gray = True
                        elif comp_key == "zero" and (not z_p and not z_a):
                            is_gray = True
                        elif comp_key == "few" and (not f_p and not f_a):
                            is_gray = True
                        elif comp_key == "two" and (not t_p and not t_a):
                            is_gray = True

                    if is_gray:
                        f1_val = None
                    else:
                        if exp_target == 1:
                            if comp_key == "zero":
                                f1_val = best_path_f1(z_p, z_a, e1_p, e1_a)
                            elif comp_key == "few":
                                f1_val = best_path_f1(f_p, f_a, e1_p, e1_a)
                            elif comp_key == "two":
                                f1_val = best_path_f1(t_p, t_a, e1_p, e1_a)
                        else:
                            if comp_key == "zero":
                                f1_val = best_path_f1(z_p, z_a, e2_p, e2_a)
                            elif comp_key == "few":
                                f1_val = best_path_f1(f_p, f_a, e2_p, e2_a)
                            elif comp_key == "two":
                                f1_val = best_path_f1(t_p, t_a, e2_p, e2_a)

                    fill_style, val_str = get_fill_and_val_str(f1_val)

                    data_cell = ws1.cell(row=current_row, column=col_num, value=val_str)
                    data_cell.fill = fill_style
                    data_cell.border = CELL_BORDER
                    data_cell.alignment = Alignment(horizontal="center", vertical="center")
                    data_cell.font = Font(name="Calibri", size=9, bold=(f1_val == 1.0 or f1_val == 0.0))

                current_row += 1

    # ---------------------------------------------------------
    # SHEET 3: F1 Heatmap (Detail - Alle Zeilen)
    # ---------------------------------------------------------
    ws2 = wb.create_sheet(title="F1 Heatmap (Detail)")
    ws2.views.sheetView[0].showGridLines = True

    ws2.row_dimensions[1].height = 100
    ws2.column_dimensions['A'].width = 36

    for i, img_id in enumerate(IMAGE_IDS):
        col_letter = get_column_letter(i + 2)
        ws2.column_dimensions[col_letter].width = 4.5
        
        cell = ws2.cell(row=1, column=i+2, value=img_id)
        cell.font = FONT_DARK_HEADER
        cell.fill = FILL_DARK_HEADER
        cell.alignment = Alignment(textRotation=90, horizontal="center", vertical="center")
        cell.border = CELL_BORDER

    cell_a1_2 = ws2.cell(row=1, column=1, value="")
    cell_a1_2.fill = FILL_DARK_HEADER
    cell_a1_2.border = CELL_BORDER

    current_row = 2

    for level_title, level_num in levels:
        ws2.row_dimensions[current_row].height = 24
        cell_sec = ws2.cell(row=current_row, column=1, value=level_title)
        cell_sec.font = FONT_SECTION
        cell_sec.fill = FILL_SECTION_BLUE
        cell_sec.alignment = Alignment(vertical="center")
        
        for c in range(2, 62):
            c_cell = ws2.cell(row=current_row, column=c)
            c_cell.fill = FILL_SECTION_BLUE
            c_cell.border = CELL_BORDER
        current_row += 1

        exp_blocks = [
            ("Experte 1", 1),
            ("Experte 2", 2)
        ]

        for exp_title, exp_target in exp_blocks:
            ws2.row_dimensions[current_row].height = 20
            cell_sub = ws2.cell(row=current_row, column=1, value=f"  {exp_title}")
            cell_sub.font = FONT_SUBSECTION
            cell_sub.fill = FILL_SUBSECTION_GREEN
            cell_sub.alignment = Alignment(vertical="center")

            for c in range(2, 62):
                c_cell = ws2.cell(row=current_row, column=c)
                c_cell.fill = FILL_SUBSECTION_GREEN
                c_cell.border = CELL_BORDER
            current_row += 1

            if exp_target == 1:
                row_specs = [
                    ("Experte 2 Präferenz → Experte 1", "exp2_pref"),
                    ("Experte 2 Alternative → Experte 1", "exp2_alt"),
                    ("KI Zero-Shot → Experte 1", "zero"),
                    ("KI Few-Shot → Experte 1", "few"),
                    ("KI Two-Stage → Experte 1", "two")
                ]
            else:
                row_specs = [
                    ("Experte 1 Präferenz → Experte 2", "exp1_pref"),
                    ("Experte 1 Alternative → Experte 2", "exp1_alt"),
                    ("KI Zero-Shot → Experte 2", "zero"),
                    ("KI Few-Shot → Experte 2", "few"),
                    ("KI Two-Stage → Experte 2", "two")
                ]

            for row_label, comp_key in row_specs:
                ws2.row_dimensions[current_row].height = 19
                lbl_cell = ws2.cell(row=current_row, column=1, value=f"  {row_label}")
                lbl_cell.font = FONT_ROW_LABEL
                lbl_cell.border = CELL_BORDER

                for img_idx, img_id in enumerate(IMAGE_IDS):
                    col_num = img_idx + 2
                    
                    e1_p_raw = parse_set(df_gt1[df_gt1["image_id"] == img_id]["praeferenz_produkt"].values[0] if len(df_gt1[df_gt1["image_id"] == img_id]) > 0 else "")
                    e1_a_raw = parse_set(df_gt1[df_gt1["image_id"] == img_id]["alternative_produkt"].values[0] if len(df_gt1[df_gt1["image_id"] == img_id]) > 0 else "")

                    e2_p_raw = parse_set(df_gt2[df_gt2["image_id"] == img_id]["praeferenz_produkt"].values[0] if len(df_gt2[df_gt2["image_id"] == img_id]) > 0 else "")
                    e2_a_raw = parse_set(df_gt2[df_gt2["image_id"] == img_id]["alternative_produkt"].values[0] if len(df_gt2[df_gt2["image_id"] == img_id]) > 0 else "")

                    z_p_raw = parse_set(df_zero[df_zero["image_id"] == img_id]["praeferenz_wundauflage"].values[0] if len(df_zero[df_zero["image_id"] == img_id]) > 0 else "")
                    z_a_raw = parse_set(df_zero[df_zero["image_id"] == img_id]["alternativ_wundauflage"].values[0] if len(df_zero[df_zero["image_id"] == img_id]) > 0 else "")

                    f_p_raw = parse_set(df_few[df_few["image_id"] == img_id]["praeferenz_wundauflage"].values[0] if len(df_few[df_few["image_id"] == img_id]) > 0 else "")
                    f_a_raw = parse_set(df_few[df_few["image_id"] == img_id]["alternativ_wundauflage"].values[0] if len(df_few[df_few["image_id"] == img_id]) > 0 else "")

                    t_p_raw = parse_set(df_two[df_two["image_id"] == img_id]["praeferenz_wundauflage"].values[0] if len(df_two[df_two["image_id"] == img_id]) > 0 else "")
                    t_a_raw = parse_set(df_two[df_two["image_id"] == img_id]["alternativ_wundauflage"].values[0] if len(df_two[df_two["image_id"] == img_id]) > 0 else "")

                    if level_num == 1:
                        e1_p, e1_a = e1_p_raw, e1_a_raw
                        e2_p, e2_a = e2_p_raw, e2_a_raw
                        z_p, z_a = z_p_raw, z_a_raw
                        f_p, f_a = f_p_raw, f_a_raw
                        t_p, t_a = t_p_raw, t_a_raw
                    elif level_num == 2:
                        e1_p, e1_a = map_level_2(e1_p_raw), map_level_2(e1_a_raw)
                        e2_p, e2_a = map_level_2(e2_p_raw), map_level_2(e2_a_raw)
                        z_p, z_a = map_level_2(z_p_raw), map_level_2(z_a_raw)
                        f_p, f_a = map_level_2(f_p_raw), map_level_2(f_a_raw)
                        t_p, t_a = map_level_2(t_p_raw), map_level_2(t_a_raw)
                    else:
                        e1_p, e1_a = map_level_3(e1_p_raw), map_level_3(e1_a_raw)
                        e2_p, e2_a = map_level_3(e2_p_raw), map_level_3(e2_a_raw)
                        z_p, z_a = map_level_3(z_p_raw), map_level_3(z_a_raw)
                        f_p, f_a = map_level_3(f_p_raw), map_level_3(f_a_raw)
                        t_p, t_a = map_level_3(t_p_raw), map_level_3(t_a_raw)

                    is_gray = False
                    if exp_target == 1:
                        if not e1_p and not e1_a:
                            is_gray = True
                        elif comp_key == "exp2_pref" and not e2_p:
                            is_gray = True
                        elif comp_key == "exp2_alt" and not e2_a:
                            is_gray = True
                        elif comp_key == "zero" and (not z_p and not z_a):
                            is_gray = True
                        elif comp_key == "few" and (not f_p and not f_a):
                            is_gray = True
                        elif comp_key == "two" and (not t_p and not t_a):
                            is_gray = True
                    else:
                        if not e2_p and not e2_a:
                            is_gray = True
                        elif comp_key == "exp1_pref" and not e1_p:
                            is_gray = True
                        elif comp_key == "exp1_alt" and not e1_a:
                            is_gray = True
                        elif comp_key == "zero" and (not z_p and not z_a):
                            is_gray = True
                        elif comp_key == "few" and (not f_p and not f_a):
                            is_gray = True
                        elif comp_key == "two" and (not t_p and not t_a):
                            is_gray = True

                    if is_gray:
                        f1_val = None
                    else:
                        if exp_target == 1:
                            if comp_key == "exp2_pref":
                                f1_val = max(set_f1(e2_p, e1_p), set_f1(e2_p, e1_a))
                            elif comp_key == "exp2_alt":
                                f1_val = max(set_f1(e2_a, e1_p), set_f1(e2_a, e1_a))
                            elif comp_key == "zero":
                                f1_val = best_path_f1(z_p, z_a, e1_p, e1_a)
                            elif comp_key == "few":
                                f1_val = best_path_f1(f_p, f_a, e1_p, e1_a)
                            elif comp_key == "two":
                                f1_val = best_path_f1(t_p, t_a, e1_p, e1_a)
                        else:
                            if comp_key == "exp1_pref":
                                f1_val = max(set_f1(e1_p, e2_p), set_f1(e1_p, e2_a))
                            elif comp_key == "exp1_alt":
                                f1_val = max(set_f1(e1_a, e2_p), set_f1(e1_a, e2_a))
                            elif comp_key == "zero":
                                f1_val = best_path_f1(z_p, z_a, e2_p, e2_a)
                            elif comp_key == "few":
                                f1_val = best_path_f1(f_p, f_a, e2_p, e2_a)
                            elif comp_key == "two":
                                f1_val = best_path_f1(t_p, t_a, e2_p, e2_a)

                    fill_style, val_str = get_fill_and_val_str(f1_val)

                    data_cell = ws2.cell(row=current_row, column=col_num, value=val_str)
                    data_cell.fill = fill_style
                    data_cell.border = CELL_BORDER
                    data_cell.alignment = Alignment(horizontal="center", vertical="center")
                    data_cell.font = Font(name="Calibri", size=9, bold=(f1_val == 1.0 or f1_val == 0.0))

                current_row += 1

    # ---------------------------------------------------------
    # SHEET 4: Produkt-Mapping (3 Ebenen)
    # ---------------------------------------------------------
    ws3 = wb.create_sheet(title="Produkt-Mapping (3 Ebenen)")
    ws3.views.sheetView[0].showGridLines = True

    all_products = set()
    for df, cols in [
        (df_gt1, ["praeferenz_produkt", "alternative_produkt"]),
        (df_gt2, ["praeferenz_produkt", "alternative_produkt"]),
        (df_zero, ["praeferenz_wundauflage", "alternativ_wundauflage"]),
        (df_few, ["praeferenz_wundauflage", "alternativ_wundauflage"]),
        (df_two, ["praeferenz_wundauflage", "alternativ_wundauflage"]),
    ]:
        for col in cols:
            for val in df[col]:
                for p in parse_set(val):
                    if p:
                        all_products.add(p.strip())

    headers_ws3 = [
        "Ebene 1: Exaktes L&R Produkt (Spezifisch)",
        "Ebene 2: L&R Unterkategorie (Produktfamilie)",
        "Ebene 3: Allgemeine Verbandsklasse (Katalog)"
    ]

    ws3.row_dimensions[1].height = 28
    ws3.append(headers_ws3)

    for c_idx in range(1, 4):
        col_letter = get_column_letter(c_idx)
        ws3.column_dimensions[col_letter].width = 42
        cell = ws3.cell(row=1, column=c_idx)
        cell.font = FONT_DARK_HEADER
        cell.fill = FILL_DARK_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = CELL_BORDER

    fill_row_even = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    fill_row_odd = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    r_num = 2
    for p_name in sorted(all_products):
        l2_set = map_level_2({p_name})
        l3_set = map_level_3({p_name})

        l2_val = list(l2_set)[0] if l2_set else p_name
        l3_val = list(l3_set)[0] if l3_set else p_name

        ws3.append([p_name, l2_val, l3_val])
        ws3.row_dimensions[r_num].height = 20

        row_fill = fill_row_even if r_num % 2 == 0 else fill_row_odd

        for c_idx in range(1, 4):
            cell = ws3.cell(row=r_num, column=c_idx)
            cell.border = CELL_BORDER
            cell.fill = row_fill
            cell.font = Font(name="Calibri", size=10)
            cell.alignment = Alignment(vertical="center", wrap_text=True)

        r_num += 1

    output_path = os.path.join(EXPORT_DIR, "Wundauflagen_F1_Heatmap_Vergleich.xlsx")
    wb.save(output_path)
    print(f"Saved: {output_path}")

if __name__ == "__main__":
    build_excel_heatmap()
