import os
import json
import ast
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from lokalisation_mapping_dictionary import map_lokalisation_explicit, EXPLICIT_LOKALISATION_RULES

BASE_DIR = "/Users/mertakdemir/Developer/uni/Modell"

def parse_val_str(val):
    if pd.isna(val) or val == "" or val == "[]" or val == "nan" or val is None or val == "?" or val == "???":
        return "keine Angabe"
    val_str = str(val).strip()
    if val_str.startswith("["):
        try:
            parsed = ast.literal_eval(val_str)
            if isinstance(parsed, list):
                return ", ".join(str(x) for x in parsed) if parsed else "keine Angabe"
        except: pass
    return val_str

def load_ki_lokalisation(sd_name):
    sd_path = os.path.join(BASE_DIR, "runs/gpt-5", sd_name)
    results = {}
    for i in range(1, 61):
        img_id = f"wunde_{i:02d}"
        bd = f"Bild{i}"
        b_path = os.path.join(sd_path, bd)
        if not os.path.exists(b_path):
            results[img_id] = "N/A"
            continue
        json_files = sorted([f for f in os.listdir(b_path) if f.startswith("run_") and f.endswith(".json")])
        if not json_files:
            results[img_id] = "N/A"
            continue
        with open(os.path.join(b_path, json_files[-1])) as f:
            data = json.load(f)
        po = data.get("parsed_output", {})
        if not po or not isinstance(po, dict):
            results[img_id] = "N/A"
            continue
        val = po.get("lokalisation")
        results[img_id] = parse_val_str(val)
    return results

def get_agreement_info(v1, v2, v3):
    valid = [v for v in [v1, v2, v3] if v and v not in ["keine Angabe", "Enthaltung / keine Angabe", "N/A"]]
    if len(valid) < 2:
        return "N/A", None
    if len(valid) == 3:
        if valid[0] == valid[1] == valid[2]:
            return "3/3 Einig", valid[0]
        if valid[0] == valid[1] or valid[0] == valid[2]:
            return "2/3 Einig", valid[0]
        if valid[1] == valid[2]:
            return "2/3 Einig", valid[1]
        return "0/3 Uneinig", None
    if len(valid) == 2:
        if valid[0] == valid[1]:
            return "2/2 Einig", valid[0]
        return "0/2 Uneinig", None
    return "0/3 Uneinig", None

def main():
    gt1_raw = pd.read_csv(os.path.join(BASE_DIR, "data/ground_truth/lohmann_rauscher/Experte1_LR_GroundTruth.csv"), sep=";")
    gt2_raw = pd.read_csv(os.path.join(BASE_DIR, "data/ground_truth/lohmann_rauscher/Experte2_LR_GroundTruth.csv"), sep=";")
    gtn_raw = pd.read_csv(os.path.join(BASE_DIR, "data/ground_truth/allgemeine_verbandsklassen.csv"))

    gt1_norm = pd.read_csv(os.path.join(BASE_DIR, "data/ground_truth/lohmann_rauscher/Experte1_LR_GroundTruth_normalised.csv"), sep=";")
    gt2_norm = pd.read_csv(os.path.join(BASE_DIR, "data/ground_truth/lohmann_rauscher/Experte2_LR_GroundTruth_normalised.csv"), sep=";")
    gtn_norm = pd.read_csv(os.path.join(BASE_DIR, "data/ground_truth/allgemeine_verbandsklassen_normalised.csv"))

    ki_zero_n = load_ki_lokalisation("zero_shot")
    ki_few_n = load_ki_lokalisation("few_shot")
    ki_two_n = load_ki_lokalisation("two_stage")

    ki_zero_lr = load_ki_lokalisation("zero_shot_lr")
    ki_few_lr = load_ki_lokalisation("few_shot_lr")
    ki_two_lr = load_ki_lokalisation("two_stage_lr")

    # Prompt example wounds to exclude from Few-Shot evaluation
    fs_lr_prompt_ex = ["wunde_04", "wunde_18"]
    fs_nurs_prompt_ex = ["wunde_18", "wunde_28"]

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Styles
    header_fill_raw = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_fill_norm = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    header_fill_ref = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
    summary_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    cell_font = Font(name="Calibri", size=10)
    bold_font = Font(name="Calibri", size=10, bold=True)
    summary_header_font = Font(name="Calibri", size=11, bold=True, color="1F4E78")

    fill_green = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid") # Expert Match / Agreement
    fill_orange = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid") # Partial Agreement
    fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9")
    )

    mapping_reference = []

    # -------------------------------------------------------------
    # TAB 1: ROHDATEN
    # -------------------------------------------------------------
    ws1 = wb.create_sheet(title="Lokalisation (Rohdaten)")
    headers1 = ["Wunde ID", "Experte 1 (L&R Raw)", "Experte 2 (L&R Raw)", "Experte 3 (NursIT Raw)"]
    ws1.append(headers1)

    for col_idx in range(1, len(headers1)+1):
        cell = ws1.cell(row=1, column=col_idx)
        cell.fill = header_fill_raw
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for i in range(1, 61):
        img_id = f"wunde_{i:02d}"
        r1 = gt1_raw[gt1_raw["image_id"] == img_id]
        r2 = gt2_raw[gt2_raw["image_id"] == img_id]
        rn = gtn_raw[gtn_raw["image_id"] == img_id]

        v1 = parse_val_str(r1["lokalisation"].values[0]) if len(r1) > 0 and "lokalisation" in r1.columns else "keine Angabe"
        v2 = parse_val_str(r2["lokalisation"].values[0]) if len(r2) > 0 and "lokalisation" in r2.columns else "keine Angabe"
        vn = parse_val_str(rn["lokalisation"].values[0]) if len(rn) > 0 and "lokalisation" in rn.columns else "keine Angabe"

        row_vals = [img_id, v1, v2, vn]
        ws1.append(row_vals)
        r_idx = ws1.max_row
        for c_idx in range(1, len(row_vals)+1):
            c = ws1.cell(row=r_idx, column=c_idx)
            c.font = cell_font
            c.border = thin_border
            c.alignment = Alignment(vertical="center", wrap_text=True)
            if c_idx == 1:
                c.alignment = Alignment(horizontal="center", vertical="center")

        if v1 != "keine Angabe": mapping_reference.append((v1, map_lokalisation_explicit(v1), "Experte 1 (L&R Raw)"))
        if v2 != "keine Angabe": mapping_reference.append((v2, map_lokalisation_explicit(v2), "Experte 2 (L&R Raw)"))
        if vn != "keine Angabe": mapping_reference.append((vn, map_lokalisation_explicit(vn), "Experte 3 (NursIT Raw)"))

    # -------------------------------------------------------------
    # TAB 2: GEMAPPT & KI
    # -------------------------------------------------------------
    ws2 = wb.create_sheet(title="Lokalisation (Gemappt & KI)")
    headers2 = [
        "Wunde ID",
        "Experte 1 (L&R Mapped)",
        "Experte 2 (L&R Mapped)",
        "Experte 3 (NursIT Mapped)",
        "Experten-Einigkeit",
        "Zero-Shot NursIT",
        "Few-Shot NursIT",
        "Two-Stage NursIT",
        "NursIT-KI-Einigkeit",
        "Zero-Shot L&R",
        "Few-Shot L&R",
        "Two-Stage L&R",
        "L&R-KI-Einigkeit"
    ]
    ws2.append(headers2)

    for col_idx in range(1, len(headers2)+1):
        cell = ws2.cell(row=1, column=col_idx)
        cell.fill = header_fill_norm
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for i in range(1, 61):
        img_id = f"wunde_{i:02d}"
        r1 = gt1_norm[gt1_norm["image_id"] == img_id]
        r2 = gt2_norm[gt2_norm["image_id"] == img_id]
        rn = gtn_norm[gtn_norm["image_id"] == img_id]

        v1_raw = parse_val_str(r1["lokalisation"].values[0]) if len(r1) > 0 and "lokalisation" in r1.columns else ""
        v2_raw = parse_val_str(r2["lokalisation"].values[0]) if len(r2) > 0 and "lokalisation" in r2.columns else ""
        vn_raw = parse_val_str(rn["lokalisation"].values[0]) if len(rn) > 0 and "lokalisation" in rn.columns else ""

        m1 = map_lokalisation_explicit(v1_raw)
        m2 = map_lokalisation_explicit(v2_raw)
        mn = map_lokalisation_explicit(vn_raw)
        exp_status, maj_val = get_agreement_info(m1, m2, mn)

        # KI NursIT
        r_zn = ki_zero_n.get(img_id, "")
        r_fn = ki_few_n.get(img_id, "")
        r_tn = ki_two_n.get(img_id, "")

        m_zn = map_lokalisation_explicit(r_zn)
        m_fn = map_lokalisation_explicit(r_fn) if img_id not in fs_nurs_prompt_ex else "Prompt-Beispiel (Nicht bewertet)"
        m_tn = map_lokalisation_explicit(r_tn)
        nursit_status = get_agreement_info(m_zn, m_fn, m_tn)[0]

        # KI L&R
        r_zlr = ki_zero_lr.get(img_id, "")
        r_flr = ki_few_lr.get(img_id, "")
        r_tlr = ki_two_lr.get(img_id, "")

        m_zlr = map_lokalisation_explicit(r_zlr)
        m_flr = map_lokalisation_explicit(r_flr) if img_id not in fs_lr_prompt_ex else "Prompt-Beispiel (Nicht bewertet)"
        m_tlr = map_lokalisation_explicit(r_tlr)
        lr_status = get_agreement_info(m_zlr, m_flr, m_tlr)[0]

        # Check Expert Matches
        hit_zn = (m_zn == mn)
        hit_fn = (m_fn == mn) if img_id not in fs_nurs_prompt_ex else False
        hit_tn = (m_tn == mn)

        hit_zlr = (m_zlr == m1 or m_zlr == m2)
        hit_flr = (m_flr == m1 or m_flr == m2) if img_id not in fs_lr_prompt_ex else False
        hit_tlr = (m_tlr == m1 or m_tlr == m2)

        if r_zn != "N/A": mapping_reference.append((r_zn, m_zn, "Zero-Shot NursIT KI"))
        if r_fn != "N/A" and img_id not in fs_nurs_prompt_ex: mapping_reference.append((r_fn, m_fn, "Few-Shot NursIT KI"))
        if r_tn != "N/A": mapping_reference.append((r_tn, m_tn, "Two-Stage NursIT KI"))
        if r_zlr != "N/A": mapping_reference.append((r_zlr, m_zlr, "Zero-Shot L&R KI"))
        if r_flr != "N/A" and img_id not in fs_lr_prompt_ex: mapping_reference.append((r_flr, m_flr, "Few-Shot L&R KI"))
        if r_tlr != "N/A": mapping_reference.append((r_tlr, m_tlr, "Two-Stage L&R KI"))

        row_vals = [
            img_id,
            m1, m2, mn, exp_status,
            m_zn, m_fn, m_tn, nursit_status,
            m_zlr, m_flr, m_tlr, lr_status
        ]

        ws2.append(row_vals)
        r_idx = ws2.max_row

        for c_idx in range(1, len(row_vals)+1):
            c = ws2.cell(row=r_idx, column=c_idx)
            c.font = cell_font
            c.border = thin_border
            c.alignment = Alignment(vertical="center", wrap_text=True)

            if c_idx == 1:
                c.alignment = Alignment(horizontal="center", vertical="center")

            # EXPERTS HIGHLIGHTING (Cols 2-5)
            elif 2 <= c_idx <= 5:
                if "3/3" in exp_status: c.fill = fill_green
                elif "2/3" in exp_status or "2/2" in exp_status: c.fill = fill_orange
                else: c.fill = fill_white
                if c_idx == 5: c.font = bold_font; c.alignment = Alignment(horizontal="center", vertical="center")

            # NURSIT KI HIGHLIGHTING
            elif c_idx == 6:
                if hit_zn: c.fill = fill_green
            elif c_idx == 7:
                if img_id in fs_nurs_prompt_ex: c.font = Font(name="Calibri", size=9, italic=True, color="7F7F7F")
                elif hit_fn: c.fill = fill_green
            elif c_idx == 8:
                if hit_tn: c.fill = fill_green
            elif c_idx == 9:
                c.alignment = Alignment(horizontal="center", vertical="center")

            # L&R KI HIGHLIGHTING
            elif c_idx == 10:
                if hit_zlr: c.fill = fill_green
            elif c_idx == 11:
                if img_id in fs_lr_prompt_ex: c.font = Font(name="Calibri", size=9, italic=True, color="7F7F7F")
                elif hit_flr: c.fill = fill_green
            elif c_idx == 12:
                if hit_tlr: c.fill = fill_green
            elif c_idx == 13:
                c.alignment = Alignment(horizontal="center", vertical="center")

    # -------------------------------------------------------------
    # ADD SUMMARY BLOCK AT THE BOTTOM
    # -------------------------------------------------------------
    ws2.append([]) # Empty row 62

    # Row 63: Summary Section Title
    ws2.append(["STATISTIK & KI-TREFFERQUOTE BEI EXPERTEN (GESAMT 60 WUNDEN, TEILEINIGKEIT & KONSENS-WUNDEN)", "", "", "", "", "", "", "", "", "", "", "", ""])
    r_hdr = ws2.max_row
    ws2.merge_cells(start_row=r_hdr, start_column=1, end_row=r_hdr, end_column=13)
    hdr_c = ws2.cell(row=r_hdr, column=1)
    hdr_c.font = summary_header_font
    hdr_c.fill = summary_fill
    hdr_c.alignment = Alignment(horizontal="left", vertical="center")

    models_config = {
        6: (ki_zero_n, [], "nursit"),
        7: (ki_few_n, fs_nurs_prompt_ex, "nursit"),
        8: (ki_two_n, [], "nursit"),
        10: (ki_zero_lr, [], "lr"),
        11: (ki_few_lr, fs_lr_prompt_ex, "lr"),
        12: (ki_two_lr, [], "lr")
    }

    # Row 64: KI Absolute Hits & Percentage across all 60 Wounds (excl. Prompt Examples for Few-Shot)
    row_64 = ["KI-Trefferquote Gesamt (vs. jeweilige Experten)", "", "", "", ""]

    for col_i in range(6, 14):
        if col_i in models_config:
            mdict, p_ex, mtype = models_config[col_i]
            hits = 0
            tot = 0
            for i in range(1, 61):
                img_id = f"wunde_{i:02d}"
                if img_id in p_ex: continue
                r1 = gt1_norm[gt1_norm["image_id"] == img_id]
                r2 = gt2_norm[gt2_norm["image_id"] == img_id]
                rn = gtn_norm[gtn_norm["image_id"] == img_id]

                m1 = map_lokalisation_explicit(parse_val_str(r1["lokalisation"].values[0])) if len(r1)>0 and "lokalisation" in r1.columns else ""
                m2 = map_lokalisation_explicit(parse_val_str(r2["lokalisation"].values[0])) if len(r2)>0 and "lokalisation" in r2.columns else ""
                mn = map_lokalisation_explicit(parse_val_str(rn["lokalisation"].values[0])) if len(rn)>0 and "lokalisation" in rn.columns else ""

                ki_mapped = map_lokalisation_explicit(mdict.get(img_id, ""))
                tot += 1
                if mtype == "nursit":
                    if ki_mapped == mn: hits += 1
                else:
                    if ki_mapped == m1 or ki_mapped == m2: hits += 1
            pct = (hits / tot) * 100
            row_64.append(f"{hits} / {tot} ({pct:.1f}%)")
        else:
            row_64.append("-")

    ws2.append(row_64)
    r64_idx = ws2.max_row
    ws2.cell(row=r64_idx, column=1).font = bold_font
    for c_idx in range(6, 14):
        c = ws2.cell(row=r64_idx, column=c_idx)
        c.font = bold_font
        c.border = thin_border
        c.alignment = Alignment(horizontal="center", vertical="center")
        if c_idx in models_config: c.fill = fill_green

    # Row 65: KI Trefferquote bei Teileinigkeit / Mehrheits-Konsens (2/3 & 3/3)
    row_65 = ["KI-Trefferquote bei Teileinigkeit / Mehrheit (2/3 & 3/3)", "", "", "", ""]

    for col_i in range(6, 14):
        if col_i in models_config:
            mdict, p_ex, mtype = models_config[col_i]
            hits = 0
            tot = 0
            for i in range(1, 61):
                img_id = f"wunde_{i:02d}"
                if img_id in p_ex: continue
                r1 = gt1_norm[gt1_norm["image_id"] == img_id]
                r2 = gt2_norm[gt2_norm["image_id"] == img_id]
                rn = gtn_norm[gtn_norm["image_id"] == img_id]

                m1 = map_lokalisation_explicit(parse_val_str(r1["lokalisation"].values[0])) if len(r1)>0 and "lokalisation" in r1.columns else ""
                m2 = map_lokalisation_explicit(parse_val_str(r2["lokalisation"].values[0])) if len(r2)>0 and "lokalisation" in r2.columns else ""
                mn = map_lokalisation_explicit(parse_val_str(rn["lokalisation"].values[0])) if len(rn)>0 and "lokalisation" in rn.columns else ""

                status, maj_val = get_agreement_info(m1, m2, mn)
                if ("3/3" in status or "2/3" in status or "2/2" in status) and maj_val:
                    tot += 1
                    ki_mapped = map_lokalisation_explicit(mdict.get(img_id, ""))
                    if ki_mapped == maj_val: hits += 1
            pct = (hits / tot) * 100 if tot > 0 else 0
            row_65.append(f"{hits} / {tot} ({pct:.1f}%)")
        else:
            row_65.append("-")

    ws2.append(row_65)
    r65_idx = ws2.max_row
    ws2.cell(row=r65_idx, column=1).font = bold_font
    for c_idx in range(6, 14):
        c = ws2.cell(row=r65_idx, column=c_idx)
        c.font = bold_font
        c.border = thin_border
        c.alignment = Alignment(horizontal="center", vertical="center")
        if c_idx in models_config: c.fill = fill_green

    # Row 66: KI Trefferquote bei 100% Einigkeit (3/3 Konsens)
    row_66 = ["KI-Trefferquote bei 100% Einigkeit (3/3 Konsens)", "", "", "", ""]

    for col_i in range(6, 14):
        if col_i in models_config:
            mdict, p_ex, mtype = models_config[col_i]
            hits = 0
            tot = 0
            for i in range(1, 61):
                img_id = f"wunde_{i:02d}"
                if img_id in p_ex: continue
                r1 = gt1_norm[gt1_norm["image_id"] == img_id]
                r2 = gt2_norm[gt2_norm["image_id"] == img_id]
                rn = gtn_norm[gtn_norm["image_id"] == img_id]

                m1 = map_lokalisation_explicit(parse_val_str(r1["lokalisation"].values[0])) if len(r1)>0 and "lokalisation" in r1.columns else ""
                m2 = map_lokalisation_explicit(parse_val_str(r2["lokalisation"].values[0])) if len(r2)>0 and "lokalisation" in r2.columns else ""
                mn = map_lokalisation_explicit(parse_val_str(rn["lokalisation"].values[0])) if len(rn)>0 and "lokalisation" in rn.columns else ""

                status, maj_val = get_agreement_info(m1, m2, mn)
                if "3/3" in status and maj_val:
                    tot += 1
                    ki_mapped = map_lokalisation_explicit(mdict.get(img_id, ""))
                    if ki_mapped == maj_val: hits += 1
            pct = (hits / tot) * 100 if tot > 0 else 0
            row_66.append(f"{hits} / {tot} ({pct:.1f}%)")
        else:
            row_66.append("-")

    ws2.append(row_66)
    r66_idx = ws2.max_row
    ws2.cell(row=r66_idx, column=1).font = bold_font
    for c_idx in range(6, 14):
        c = ws2.cell(row=r66_idx, column=c_idx)
        c.font = bold_font
        c.border = thin_border
        c.alignment = Alignment(horizontal="center", vertical="center")
        if c_idx in models_config: c.fill = fill_green

    for r_idx in range(r64_idx, r66_idx + 1):
        for c_idx in range(1, 6):
            ws2.cell(row=r_idx, column=c_idx).border = thin_border

    # -------------------------------------------------------------
    # TAB 3: MAPPING-TABELLE (REFERENZ)
    # -------------------------------------------------------------
    ws3 = wb.create_sheet(title="Mapping-Tabelle (Referenz)")
    headers3 = ["Ziel-Kategorie (Körperregion)", "Roh-Eingabe / Originaler Freitext", "Herkunft", "Mapping-Regel / Status"]
    ws3.append(headers3)

    for col_idx in range(1, len(headers3)+1):
        cell = ws3.cell(row=1, column=col_idx)
        cell.fill = header_fill_ref
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    unique_ref = {}
    for raw_str, mapped_str, src in mapping_reference:
        if raw_str not in unique_ref:
            unique_ref[raw_str] = (mapped_str, src)

    standard_categories = {
        "Fuß",
        "Bein",
        "Gesäß / Sakral",
        "Arm / Hand",
        "Abdomen"
    }

    sorted_ref = sorted(unique_ref.items(), key=lambda x: (x[1][0], x[0]))

    for raw_str, (mapped_str, src) in sorted_ref:
        if mapped_str == "Enthaltung / keine Angabe":
            status_str = "Enthaltung / keine Angabe"
        elif raw_str in EXPLICIT_LOKALISATION_RULES:
            status_str = "1:1 Wörterbuch-Regel (Hauptkategorie)"
        elif mapped_str in standard_categories:
            status_str = "Textbasierte Regel (Körperregion)"
        else:
            status_str = "Freitext belassen (Kein Matching)"

        row_vals = [mapped_str, raw_str, src, status_str]
        ws3.append(row_vals)
        r_idx = ws3.max_row

        for c_idx in range(1, len(row_vals)+1):
            c = ws3.cell(row=r_idx, column=c_idx)
            c.font = cell_font
            c.border = thin_border
            c.alignment = Alignment(vertical="center", wrap_text=True)
            if c_idx == 1: c.font = bold_font
            if c_idx == 4:
                c.alignment = Alignment(horizontal="center", vertical="center")
                if "Freitext" in status_str:
                    c.fill = fill_orange

    for ws in [ws1, ws2, ws3]:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or "")
                max_len = max(max_len, len(val_str))
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 14), 52)

    os.makedirs(os.path.join(BASE_DIR, "exports"), exist_ok=True)
    output_file = os.path.join(BASE_DIR, "exports/Lokalisation_Vergleich_Experten_KI.xlsx")
    wb.save(output_file)
    print(f"Erfolgreich aktualisiert: {output_file}")

if __name__ == "__main__":
    main()
